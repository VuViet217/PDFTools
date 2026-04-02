"""
API Router - PDF Table to Excel conversion
"""
import io
import logging
import tempfile
import os
import time
from fastapi import APIRouter, UploadFile, File, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)

router = APIRouter()


def remove_file(path: str):
    """Background task to remove file after download"""
    try:
        time.sleep(5)
        if os.path.exists(path):
            os.remove(path)
    except:
        pass


def extract_tables_from_pdf(pdf_bytes: bytes) -> list:
    """
    Extract all tables from all pages of a PDF file.
    Returns list of dicts: [{'page': 1, 'table_index': 0, 'data': [[...], ...]}, ...]
    """
    tables = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            page_tables = page.extract_tables()
            if page_tables:
                for table_idx, table in enumerate(page_tables):
                    if table and any(any(cell for cell in row) for row in table):
                        # Clean up None values
                        clean_table = []
                        for row in table:
                            clean_row = [(cell or "").strip() for cell in row]
                            clean_table.append(clean_row)
                        tables.append({
                            'page': page_idx + 1,
                            'table_index': table_idx,
                            'data': clean_table
                        })
    return tables


def tables_to_excel(tables: list) -> bytes:
    """
    Convert extracted tables to an Excel file.
    Each table gets its own sheet or they are placed sequentially in one sheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tables"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(vertical='center', wrap_text=True)

    current_row = 1

    for t_idx, table_info in enumerate(tables):
        page = table_info['page']
        data = table_info['data']

        if not data:
            continue

        # Table header label
        label_cell = ws.cell(row=current_row, column=1,
                             value=f"Page {page} - Table {table_info['table_index'] + 1}")
        label_cell.font = Font(bold=True, size=12, color="2F5496")
        current_row += 1

        for row_idx, row in enumerate(data):
            for col_idx, cell_value in enumerate(row):
                cell = ws.cell(row=current_row, column=col_idx + 1, value=cell_value)
                cell.border = thin_border

                if row_idx == 0:
                    # First row as header
                    cell.font = header_font_white
                    cell.fill = header_fill
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

            current_row += 1

        # Auto-width columns
        for col_idx in range(len(data[0]) if data else 0):
            max_len = 0
            col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
            for row in data:
                if col_idx < len(row):
                    max_len = max(max_len, len(str(row[col_idx] or "")))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)

        # Gap between tables
        current_row += 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.post('/pdf-to-excel')
async def pdf_to_excel(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = None
):
    """Convert PDF tables to Excel file"""
    try:
        if not file.filename.lower().endswith('.pdf'):
            raise HTTPException(status_code=400, detail='Only PDF files are supported')

        contents = await file.read()
        if len(contents) > 50 * 1024 * 1024:
            raise HTTPException(status_code=413, detail='File too large (max 50MB)')

        # Extract tables
        tables = extract_tables_from_pdf(contents)

        if not tables:
            raise HTTPException(
                status_code=400,
                detail='No tables found in this PDF. Scanned PDFs may not be supported.'
            )

        # Convert to Excel
        excel_bytes = tables_to_excel(tables)

        # Save to temp file
        base_name = os.path.splitext(file.filename)[0]
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        with os.fdopen(tmp_fd, 'wb') as f:
            f.write(excel_bytes)

        if background_tasks:
            background_tasks.add_task(remove_file, tmp_path)

        return FileResponse(
            path=tmp_path,
            filename=f"{base_name}_tables.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error converting PDF to Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post('/pdf-to-excel/preview')
async def pdf_to_excel_preview(
    file: UploadFile = File(...)
):
    """Preview tables found in PDF (returns JSON)"""
    try:
        if not file.filename.lower().endswith('.pdf'):
            raise HTTPException(status_code=400, detail='Only PDF files are supported')

        contents = await file.read()
        if len(contents) > 50 * 1024 * 1024:
            raise HTTPException(status_code=413, detail='File too large (max 50MB)')

        tables = extract_tables_from_pdf(contents)

        return {
            'total_tables': len(tables),
            'tables': [
                {
                    'page': t['page'],
                    'table_index': t['table_index'],
                    'rows': len(t['data']),
                    'cols': len(t['data'][0]) if t['data'] else 0,
                    'preview': t['data'][:5]  # First 5 rows preview
                }
                for t in tables
            ]
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error previewing PDF tables: {e}")
        raise HTTPException(status_code=500, detail=str(e))
