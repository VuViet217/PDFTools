from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
import io
from typing import Dict, List, Any

router = APIRouter()

def read_excel_file(file_content: bytes) -> Dict[str, Any]:
    """
    Read Excel file and extract data from the first sheet
    """
    try:
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        
        return {
            'data': data,
            'sheet_name': ws.title,
            'dimensions': ws.dimensions
        }
    except Exception as e:
        raise Exception(f"Failed to read Excel file: {str(e)}")

def compare_excel_data(data1: List[List[Any]], data2: List[List[Any]]) -> Dict[str, Any]:
    """
    Compare two Excel datasets row by row
    Returns comparison result with changes
    """
    comparison_rows = []
    stats = {
        'unchanged': 0,
        'added': 0,
        'removed': 0,
        'modified': 0
    }
    
    # Process rows by comparing line by line
    max_rows = max(len(data1), len(data2))
    
    for idx in range(max_rows):
        row1 = data1[idx] if idx < len(data1) else None
        row2 = data2[idx] if idx < len(data2) else None
        
        if row1 is None:
            # Row added in file2
            comparison_rows.append({
                'status': 'added',
                'changes': None,
                'file1': None,
                'file2': row2
            })
            stats['added'] += 1
        elif row2 is None:
            # Row removed from file1
            comparison_rows.append({
                'status': 'removed',
                'changes': None,
                'file1': row1,
                'file2': None
            })
            stats['removed'] += 1
        else:
            # Both rows exist - check for changes
            row_changes = {}
            max_cols = max(len(row1), len(row2))
            
            for col_idx in range(max_cols):
                cell1 = row1[col_idx] if col_idx < len(row1) else None
                cell2 = row2[col_idx] if col_idx < len(row2) else None
                
                col_letter = get_column_letter(col_idx + 1)
                
                if cell1 != cell2:
                    if cell1 is None:
                        row_changes[col_letter] = {
                            'status': 'added',
                            'file1': '',
                            'file2': str(cell2) if cell2 is not None else ''
                        }
                    elif cell2 is None:
                        row_changes[col_letter] = {
                            'status': 'removed',
                            'file1': str(cell1) if cell1 is not None else '',
                            'file2': ''
                        }
                    else:
                        row_changes[col_letter] = {
                            'status': 'modified',
                            'file1': str(cell1) if cell1 is not None else '',
                            'file2': str(cell2) if cell2 is not None else ''
                        }
            
            if row_changes:
                comparison_rows.append({
                    'status': 'modified',
                    'changes': row_changes,
                    'file1': row1,
                    'file2': row2
                })
                stats['modified'] += 1
            else:
                comparison_rows.append({
                    'status': 'unchanged',
                    'changes': None,
                    'file1': row1,
                    'file2': row2
                })
                stats['unchanged'] += 1
    
    return {
        'rows': comparison_rows,
        'stats': stats
    }

def get_column_letter(col_num: int) -> str:
    """Convert column number to letter (1 -> A, 27 -> AA)"""
    col_letter = ''
    while col_num > 0:
        col_num -= 1
        col_letter = chr(65 + (col_num % 26)) + col_letter
        col_num //= 26
    return col_letter

@router.post("/excel-compare")
async def compare_excel(file1: UploadFile = File(...), file2: UploadFile = File(...)):
    """
    API endpoint to compare two Excel files
    """
    try:
        # Validate file types
        if not file1.filename or not file1.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File 1 must be an Excel file (.xlsx or .xls)")
        
        if not file2.filename or not file2.filename.lower().endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File 2 must be an Excel file (.xlsx or .xls)")
        
        # Read files
        file1_content = await file1.read()
        file2_content = await file2.read()
        
        # Parse Excel files
        data1 = read_excel_file(file1_content)
        data2 = read_excel_file(file2_content)
        
        # Compare data
        comparison = compare_excel_data(data1['data'], data2['data'])
        
        return JSONResponse({
            'file1': file1.filename,
            'file2': file2.filename,
            'sheet1': data1['sheet_name'],
            'sheet2': data2['sheet_name'],
            'comparison': comparison,
            'stats': comparison['stats']
        })
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
